import openpyxl

def processStudent(pathToXlsx, year, pathToProcessed):
    wrkbk = openpyxl.load_workbook(pathToXlsx)
    wrkbk = wrkbk[year]

    filepath = pathToProcessed
    wb = openpyxl.Workbook()
    wb.save(filepath)

    wb = openpyxl.load_workbook(pathToProcessed)
    sheet = wb.active
    #print(wrkbk.max_row)

    newCellRow = 1

    for j in range(1, wrkbk.max_row + 1):
        lvl = wrkbk.cell(row=j, column=2).value
        if lvl is not None:
            #edLvl = edLvlRaw #edLvlRaw.rsplit('; ', 1)[-1]
            #print(edLvl)
            cell_obj = wrkbk.cell(row=j, column=5)
            txt = cell_obj.value
            x = ""
            x2 = ""
            if lvl == "Student":
                if txt is not None:
                    x = txt.split("; ")
                    other_cell_obj = wrkbk.cell(row=j, column=6)
                    txt2 = other_cell_obj.value
                    if txt2 is not None:
                        x2 = txt2.split("; ")
                        for i in x:
                            for jj in x2:
                                if newCellRow < 1048576:
                                    # print(i + jj)
                                    sheet.cell(row=newCellRow, column=1).value = i
                                    sheet.cell(row=newCellRow, column=2).value = jj
                                    newCellRow = newCellRow + 1


    wb.save(filepath)

def processOther(pathToXlsx, year, pathToProcessed):
    wrkbk = openpyxl.load_workbook(pathToXlsx)
    wrkbk = wrkbk[year]

    filepath = pathToProcessed
    wb = openpyxl.Workbook()
    wb.save(filepath)

    wb = openpyxl.load_workbook(pathToProcessed)
    sheet = wb.active
    #print(wrkbk.max_row)

    newCellRow = 1

    for j in range(1, wrkbk.max_row + 1):
        lvl = wrkbk.cell(row=j, column=2).value
        if lvl is not None:
            #edLvl = edLvlRaw #edLvlRaw.rsplit('; ', 1)[-1]
            #print(edLvl)
            cell_obj = wrkbk.cell(row=j, column=5)
            txt = cell_obj.value
            x = ""
            x2 = ""
            if lvl == "other":
                if txt is not None:
                    x = txt.split("; ")
                    other_cell_obj = wrkbk.cell(row=j, column=6)
                    txt2 = other_cell_obj.value
                    if txt2 is not None:
                        x2 = txt2.split("; ")
                        for i in x:
                            for jj in x2:
                                if newCellRow < 1048576:
                                    # print(i + jj)
                                    sheet.cell(row=newCellRow, column=1).value = i
                                    sheet.cell(row=newCellRow, column=2).value = jj
                                    newCellRow = newCellRow + 1


    wb.save(filepath)

def processElse(pathToXlsx, year, pathToProcessed):
    wrkbk = openpyxl.load_workbook(pathToXlsx)
    wrkbk = wrkbk[year]

    filepath = pathToProcessed
    wb = openpyxl.Workbook()
    wb.save(filepath)

    wb = openpyxl.load_workbook(pathToProcessed)
    sheet = wb.active
    #print(wrkbk.max_row)

    newCellRow = 1

    for j in range(1, wrkbk.max_row + 1):
        lvl = wrkbk.cell(row=j, column=2).value
        if lvl is not None:
            #edLvl = edLvlRaw #edLvlRaw.rsplit('; ', 1)[-1]
            #print(edLvl)
            cell_obj = wrkbk.cell(row=j, column=5)
            txt = cell_obj.value
            x = ""
            x2 = ""
            if ((lvl != "other") and (lvl != "Student")):
                if txt is not None:
                    x = txt.split("; ")
                    other_cell_obj = wrkbk.cell(row=j, column=6)
                    txt2 = other_cell_obj.value
                    if txt2 is not None:
                        x2 = txt2.split("; ")
                        for i in x:
                            for jj in x2:
                                if newCellRow < 1048576:
                                    # print(i + jj)
                                    sheet.cell(row=newCellRow, column=1).value = i
                                    sheet.cell(row=newCellRow, column=2).value = jj
                                    newCellRow = newCellRow + 1


    wb.save(filepath)

#processStudent("Clean_survey_data.xlsx", '2016 Survey Data', "../processed/2016/profession/byProfStudent.xlsx")
processOther("Clean_survey_data.xlsx", '2016 Survey Data', "../processed/2016/profession/byProfOther.xlsx")
processElse("Clean_survey_data.xlsx", '2016 Survey Data', "../processed/2016/profession/byProfElse.xlsx")