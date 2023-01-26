import openpyxl

def processByCountry(pathToXlsx, year, pathToProcessed, countryy):
    wrkbk = openpyxl.load_workbook(pathToXlsx)
    wrkbk = wrkbk[year]

    filepath = pathToProcessed
    wb = openpyxl.Workbook()
    wb.save(filepath)

    wb = openpyxl.load_workbook(pathToProcessed)
    sheet = wb.active

    newCellRow = 1

    for j in range(1, wrkbk.max_row + 1):
        country = wrkbk.cell(row=j, column=1).value
        cell_obj = wrkbk.cell(row=j, column=6)
        txt = cell_obj.value
        x = ""
        x2 = ""
        if country == countryy:
            if txt != "NA":
                x = txt.split(";")
                other_cell_obj = wrkbk.cell(row=j, column=7)
                txt2 = other_cell_obj.value
                if txt2 != "NA":
                    x2 = txt2.split(";")
                    for i in x:
                        for jj in x2:
                            print(i + jj)
                            sheet.cell(row=newCellRow, column=1).value = i
                            sheet.cell(row=newCellRow, column=2).value = jj
                            newCellRow = newCellRow + 1


    wb.save(filepath)

processByCountry("Clean_survey_data.xlsx", '2018 Survey Data', "../processed/byCountryUK.xlsx", "United Kingdom")