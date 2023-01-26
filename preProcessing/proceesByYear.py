import openpyxl

def process2016(pathToXlsx, year, pathToProcessed):
    wrkbk = openpyxl.load_workbook(pathToXlsx)
    wrkbk = wrkbk[year]

    filepath = pathToProcessed
    wb = openpyxl.Workbook()
    wb.save(filepath)

    wb = openpyxl.load_workbook(pathToProcessed)
    sheet = wb.active

    newCellRow = 1

    for j in range(1, wrkbk.max_row + 1):
        cell_obj = wrkbk.cell(row=j, column=5)
        txt = cell_obj.value
        x = ""
        x2 = ""
        if txt is not None:
            x = txt.split("; ")
            other_cell_obj = wrkbk.cell(row=j, column=6)
            txt2 = other_cell_obj.value
            if txt2 is not None:
                x2 = txt2.split("; ")
                for i in x:
                    for jj in x2:
                        if newCellRow < 1048576:
                            sheet.cell(row=newCellRow, column=1).value = i
                            sheet.cell(row=newCellRow, column=2).value = jj
                            newCellRow = newCellRow + 1


    wb.save(filepath)

def process2018(pathToXlsx, year, pathToProcessed):
    wrkbk = openpyxl.load_workbook(pathToXlsx)
    wrkbk = wrkbk[year]

    filepath = pathToProcessed
    wb = openpyxl.Workbook()
    wb.save(filepath)

    wb = openpyxl.load_workbook(pathToProcessed)
    sheet = wb.active

    newCellRow = 1

    for j in range(1, wrkbk.max_row + 1):
        cell_obj = wrkbk.cell(row=j, column=6)
        txt = cell_obj.value
        x = ""
        x2 = ""
        if txt != "NA":
            x = txt.split(";")
            other_cell_obj = wrkbk.cell(row=j, column=7)
            txt2 = other_cell_obj.value
            if txt2 != "NA":
                x2 = txt2.split(";")
                for i in x:
                    for jj in x2:
                        if newCellRow < 1048576:
                            sheet.cell(row=newCellRow, column=1).value = i
                            sheet.cell(row=newCellRow, column=2).value = jj
                            newCellRow = newCellRow + 1


    wb.save(filepath)

#process2016("Clean_survey_data.xlsx", '2016 Survey Data', "../processed/all/2016.xlsx")
process2018("Clean_survey_data.xlsx", '2018 Survey Data', "../processed/all/2018.xlsx")