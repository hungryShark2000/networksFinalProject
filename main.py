# import module
import openpyxl

# load excel with its path
wrkbk = openpyxl.load_workbook("preProcessing/Clean_survey_data.xlsx")
wrkbk = wrkbk['2016 Survey Data']

filepath = "processed/processedTest.xlsx"
wb = openpyxl.Workbook()
wb.save(filepath)

wb = openpyxl.load_workbook("processed/processedTest.xlsx")
sheet=wb.active

newCellRow = 1

for j in range(2, 100):#sh.max_row + 1):
    cell_obj = wrkbk.cell(row=j, column=5)
    txt = cell_obj.value
    x = ""
    x2 = ""
    if txt is not None:
        x = txt.split("; ")
        other_cell_obj = wrkbk.cell(row=j, column=6)
        txt2 = other_cell_obj.value
        if txt2 is not None:
            x2 = txt2.split("; ")
            for i in x:
                for jj in x2:
                    print(i + jj)
                    sheet.cell(row=newCellRow, column=1).value = i
                    sheet.cell(row=newCellRow, column=2).value = jj
                    newCellRow = newCellRow + 1

    print(x)
wb.save(filepath)

